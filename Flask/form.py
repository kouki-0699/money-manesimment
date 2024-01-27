from flask import Flask, request, render_template
from openpyxl import Workbook, load_workbook  # load_workbook を追加

app = Flask(__name__)

# エクセルファイル名
excel_file = 'allowance.xlsx'

@app.route('/')
def index():
    return render_template('maney.html')

import gspread  
from oauth2client.service_account import ServiceAccountCredentials

@app.route('/process_data', methods=['POST'])
def process_data():
    month = request.form['month']
    date =request.form['date']  
    income = request.form['income']
    expenses = request.form['expenses']
    free_text = request.form['free_text']
    
    # データをエクセルファイルに書き込む
    write_to_excel(date, month, income, expenses ,free_text )
    
    return render_template('index.html')

def write_to_excel(date, item, amount):
    wb = None
    try:
        # エクセルファイルを開くか新規作成
        try:
            wb = load_workbook(excel_file)  # 既存のワークブックを読み込む
            ws = wb.active
        except:
            wb = Workbook()
            ws = wb.active

        # データを追加
        ws.append([date, item, amount])

        # ファイル保存
        wb.save(excel_file)
    finally:
        if wb:
            wb.close()

if __name__ == '__main__':
    app.run(debug=True)
